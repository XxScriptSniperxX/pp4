# coding: UTF-8
"""
File regrouping objects and functions related to exports, pdf, excel, ...
"""
if __name__ == '__main__':
    from addDll.dll import addPATH
    addPATH()

import os, sys
sys.path.append(os.path.expandvars(r"$GREENCZ\scriptsResources\Lib\site-packages"))
import openpyxl
from PIL import Image
import typing
from PySide6 import QtWidgets, QtGui, QtCore
from Utils.PluginGui.resources.progress_bar import ProgressBar

# uncomment following line for standalone use of the plugin
# app = QtWidgets.QApplication().instance()
default_progress_bar_placeholder = ProgressBar()


def pdfExport(plots_to_export_list:list, report_details_page, progress_bar=default_progress_bar_placeholder,
              export_file_path:str="DEFAULT_EXPORT_FILENAME.pdf"):
    """
    This function returns the list of PIL.Images to save to the target pdf file dest.

    :param plots_to_export_list:
    :param progress_bar:
    :param export_file_path:
    """
    progress_bar.showProgressBar()
    # plots_to_export_list = [graphPage.plot_widget for graphPage in report_manager.graph_manager.graph_builder.values() if graphPage.checkState > 0]
    img_list = []
    filename_without_ext = os.path.splitext(os.path.basename(export_file_path))[0]
    output_dir = os.path.dirname(export_file_path)
    length = len(plots_to_export_list)
    # exporting plotwidgets to img
    for ind, plot in enumerate(plots_to_export_list):
        progress_bar.updateProgressBar(int(((ind + 1) / length) * 100), f"Page {ind + 1} out of {length}")
        pixmap = plot.toPixmap(QtCore.QSize(1200 * 4, 900 * 4), 4)
        img_path = os.path.join(output_dir, filename_without_ext + f'_page_{ind}.png')
        pixmap.save(img_path, format="PNG")
        img_list.append(Image.open(img_path))
    # exporting the information panel to img
    img_list.append(Image.open(report_details_page.toPixmap(output_dir, filename_without_ext)))
    img_list[0].save(os.path.join(output_dir, filename_without_ext + ".pdf"), append_images=img_list[1:], save_all=True)

def excelExport(graph_manager, export_file_path:str="DEFAULT_EXPORT_FILENAME.xlsx"):
    """
    Main function to export a report to Excel format.

    :param graph_manager: graph_creator.GraphManager object, instance created via report_manager.py.
    :param export_file_path: file where final file will be saved.
    """
    ef = ExcelFormatter(graph_manager, export_file_path)
    ef.setPage(ef.first_page_title)
    ef.initCriteriaCells()
    ef.populateCriteriaCells()
    ef.setTargetGraphType("twoDimGraph")
    ef.setPage(ef.second_page_title)
    ef.initDataCells()
    ef.populateDataCells()
    ef.saveWorkbook()

class ExcelFormatter():
    """
    Class setting up data for excel export format and responsible for creating and saving
    the openpyxl.Workbook() at target location.
    The workbook is filled row by row for the criteria page:
    row n - 1 : column 1            |   column 2           | ... |  column p - 2         |
    row n     : GraphPage n's title |
        n + 1 :                     | Taskpad x's name     | ... | Taskpad p's name      |
        n + 2 : criteria 1 title    | Value for taskpad 1  | ... | Value for taskpad p   |
        .
        .
        .
        n + m : criteria m-2 title  | Value for taskpad m-2| ... | Value for taskpad p  |
    Then, the wb is filled column-wise for the "data" page.
    """

    def __init__(self, graph_manager, filePath):
        self.wb = openpyxl.Workbook()
        self.first_page_title = "Criteria"
        self.second_page_title = "Data"
        self.target_subgraph_type = "1DHorizontalCriteria"
        self.wb.active.title = self.first_page_title
        self.wb.create_sheet(index=1, title=self.second_page_title)
        self.sh = None
        self.graph_manager = graph_manager
        self.current_row: int = 1
        self.current_column: int = 1
        self.file_path = filePath
        self.taskpad_col_index_dict = None
        self.taskpad_col_index_dict = self.setupTaskpadColIndDict()

    def initCriteriaCells(self):
        """With Openpyxl, cells have to be initialized to be accessed.
        sh.cell(row=maxRow, column=maxCol) will enable all cells from "A1" to this specific cell."""
        sh = self.wb.get_sheet_by_name(self.first_page_title)
        graphpage_title_row_offset = 2
        subgraph_next_graphpage_rowgap = 3
        maxRow = self.graph_manager.getMaxGraphpageIndex() * \
                 (graphpage_title_row_offset + subgraph_next_graphpage_rowgap) \
                 + self.graph_manager.getMaxSubgraphIndex()
        maxCol = 1 + self.graph_manager.getMaxTaskpadIndex()
        border_cell = sh.cell(row=maxRow, column=maxCol)
        # A1 : row = 1, col = 1

    def oneDimGraphGenerator(self):
        """Function yielding the subgraphs of the target type."""
        for graphPage in self.graph_manager.graph_builder.values():
            for subgraph in graphPage.sub_Graphs.values():
                if subgraph.sub_graph_type == self.target_subgraph_type:
                    yield subgraph

    def oneDimTaskpadNameGenerator(self, subgraph):
        """
        Function yielding the taskpad names (i.e. the vehicule name) for the given subgraph in input.

        :param subgraph: graph_creator.SubGraph
        """
        crit_key = list(subgraph.criteria_values_dict)[0]
        for tup in subgraph.criteria_values_dict[crit_key]:
            yield(tup[1])

    def setTitleCell(self, title, column_index=1):
        """
        Sets the cell @ 'A + self.current_row' to the title value.

        :param title: str
        """
        self.sh.cell(row=self.current_row, column=column_index, value=title)
        # example : A1 cell is @ row = 1 and col = 1

    def offsetCurrentRow(self, value:int):
        """
        Offsets the value of self.current_row

        :param value: int, value of the offset to apply
        """
        self.current_row += value

    def setCurrentRow(self, value: int):
        """ Sets the value of current row to the input value """
        if value > 0 :
            self.current_row = value
        else:
            raise ValueError(f"input must be equal to 1 or higher, value input : {value}")

    def setTaskpadNames(self, subgraph):
        """
        Sets, for the current row, the name of the taskpad (or vehicule,
        depending on renaming made in SysA) in the criteria sheet.

        :param subgraph: subgraph containing at least one
        :return:
        """
        sh = self.wb.get_sheet_by_name(self.first_page_title)
        col = 2
        # example : 'A1' cell is @ row = 1 and col = 1
        for vehicule_name in self.oneDimTaskpadNameGenerator(subgraph):
            sh.cell(row=self.current_row, column=col, value=vehicule_name)
            col += 1

    def graphPageWithTargetSubgraphGenerator(self):
        """
        This function yields a graph page if it contains a surbgraph of type self.target_subgraph_type
        :return: graph_creator.GraphPage object
        """
        for graphPage in self.graph_manager.graph_builder.values():
            if any(subgraph.sub_graph_type == self.target_subgraph_type for subgraph in graphPage.sub_Graphs.values()):
                yield graphPage

    def populateGraphPagesRelatedCells(self):
        """Sets up the titles cells for its subgraphs criterias. Then calls the subgraph cell filling routine."""
        for page in self.graphPageWithTargetSubgraphGenerator():
            self.setTitleCell(page.title)
            self.offsetCurrentRow(2)
            self.setTaskpadNames(list(page.sub_Graphs.values())[0])
            self.offsetCurrentRow(1)
            self.populateSubgraphRelatedCells(page.sub_Graphs.values())
            self.offsetCurrentRow(3)

    def populateSubgraphRelatedCells(self, subgraph_list):
        """
        For each subgraph in the list from input, prints the criteria data informations such as follows:
        criteria name -> value for vehicule first vehicule, ..., value for vehicule n

        :param subgraph_list: list of graph_creator.SubGraphs
        """
        sh = self.wb.get_sheet_by_name(self.first_page_title)
        for subgraph in subgraph_list:
            if subgraph.sub_graph_type == self.target_subgraph_type:
                title = subgraph.sub_graph_title
                self.setTitleCell(title)
                val_name_list = list(subgraph.criteria_values_dict.values())[0]
                for tuple_value_name in val_name_list:
                    # There should only be one list of value names in a subgraph criteria_values_dict
                    sh.cell(row=self.current_row,
                            column=self.taskpad_col_index_dict[tuple_value_name[1]],
                            value=tuple_value_name[0][0])
                self.offsetCurrentRow(1)

    def populateCriteriaCells(self):
        """Function to call to fill the workbook"""
        self.populateGraphPagesRelatedCells()

    def saveWorkbook(self):
        """Saves the workbook at self.file_path location (instance input)"""
        self.wb.save(self.file_path)

    def setupTaskpadColIndDict(self):
        """Creates a dict for each taskpad name(i.e. Vehicule name) as key, and their associated column in the workbook
        in order to ease value and taskpad association for populateSubgraphRelatedCells function."""
        taskpad_index_dict = {}
        column_index = 2
        for graphpage in self.graphPageWithTargetSubgraphGenerator():
            first_subgraph = list(graphpage.sub_Graphs.values())[0]
            first_criteria_list = list(first_subgraph.criteria_values_dict.values())[0]
            for index_col, tuple_value_taskpadname in enumerate(first_criteria_list):
                taskpad_index_dict[tuple_value_taskpadname[1]] = index_col + column_index
            return taskpad_index_dict

    def setPage(self, page_name: str):
        """ Sets the page to Data """
        self.sh = self.wb.get_sheet_by_name(page_name)

    def setTargetGraphType(self, target_type_name:str):
        self.target_subgraph_type = target_type_name

    def resetCurrentRow(self):
        self.current_row = 1

    def setCurrentColumn(self, index: int):
        """ Sets the column index to the input provided """
        if index > 0:
            self.current_column = index
        else:
            raise ValueError(f"input must be equal to 1 or higher, value input : {index}")

    def dataExportSubgraphsGenerator(self):
        """ Generator yielding the subgraphs matching the target subgraph type and that have the
        data export excel option to True (info gathered in the templated instanciating the plugin) """
        for graph_page in self.graph_manager.graph_builder.values():
            for subgraph in graph_page.sub_Graphs.values():
                if subgraph.sub_graph_type == self.target_subgraph_type:
                    if subgraph.data_export_excel:
                        yield subgraph

    def getTotalNumberOfDataExportSubgraphs(self):
        """ Returns the total number of subgraphs to be exported to the Data page """
        return sum(1 for subgraph in self.dataExportSubgraphsGenerator())

    def getMaxColumn(self):
        """ Returns the largest index number for columns. A maximum of 9 taskpad columns (i.e. vehicule names) are
        displayed between two consecutive variables in the provided example file.
        The same behavior is implemented to keep the clients' excel post processing scripts compatible. """
        return 9 * (1 + self.getTotalNumberOfDataExportSubgraphs())

    def getMaxRowFromVariables(self):
        """ Returns the max length of the lists of variables in the target subgraphs. """
        maxLength = 0
        for subgraph in self.dataExportSubgraphsGenerator():
            for taskpad in subgraph.two_dim_curve_dict.values():
                for var_dict in taskpad.values():
                    if len(var_dict["values"]) > maxLength:
                        maxLength = len(var_dict["values"])
        return maxLength

    def initDataCells(self):
        """ Caching the cells on the 'Data' page . """
        data_title_rows_index = 2
        maxRow = self.getMaxRowFromVariables() + data_title_rows_index
        maxColumn = self.getMaxColumn()
        self.sh.cell(row=maxRow, column=maxColumn)


    def setupDataXAxisColumns(self):
        """ Sets the data in the first 9 columns for Data page, which refer to the xAxis data of
        the 9 (or less) potential graphs from the report. The first subgraph will be used as reference"""
        if self.getTotalNumberOfDataExportSubgraphs() > 0:
            ref_subgraph = next(self.dataExportSubgraphsGenerator())
            x_axis_title = ref_subgraph.x_axis_title
            y_axis_title = ref_subgraph.y_axis_title
            # x_title_row = 1
            # taskpad_title_row = 2
            self.resetCurrentRow()
            self.setTitleCell(x_axis_title)
            for index, (taskpad_name, taskpad_variables_dicts) in enumerate(ref_subgraph.two_dim_curve_spline_criteria_dict.items()):
                self.setCurrentRow(2)
                self.setCurrentColumn(index + 1)
                self.setTitleCell(taskpad_name, column_index=self.current_column)
                self.offsetCurrentRow(1)
                for var_id, var_dict in taskpad_variables_dicts.items():
                   if var_dict["item"] == "x_axis":
                       for row_index, row_value in enumerate(var_dict["values"]):
                           row = self.current_row + row_index
                           self.sh.cell(row, self.current_column, row_value)

    def setupDataYAxisColumns(self):
        self.resetCurrentRow()
        self.resetCurrentRow()
        for subgraph_index, subgraph in enumerate(self.dataExportSubgraphsGenerator()):
            self.setCurrentRow(1)
            var_title = subgraph.y_axis_title
            # first var of first subgraph is displayed starting at 10th column
            subgraph_column_start = 1 + (subgraph_index + 1) * 9
            self.setCurrentColumn(subgraph_column_start)
            self.setTitleCell(var_title, column_index=subgraph_column_start)
            for local_index, (taskpad_name, taskpad_variables_dicts) in enumerate(subgraph.two_dim_curve_spline_criteria_dict.items()):
                # 9 taskpad max displayed in data page.
                if local_index > 8:
                    continue
                self.setCurrentRow(2)
                taskpad_col_index = self.current_column + local_index
                self.setTitleCell(taskpad_name, column_index=taskpad_col_index)
                self.offsetCurrentRow(1)
                for var_id, var_dict in taskpad_variables_dicts.items():
                    if var_dict["item"] == "y_axis":
                        for row_index, row_value in enumerate(var_dict["values"]):
                            row = self.current_row + row_index
                            self.sh.cell(row, taskpad_col_index, row_value)


    def populateDataCells(self):
        """ Populates the 'Data' page of the excel file with the yAxis data from all taskpads for each subgraph """
        self.setupDataXAxisColumns()
        self.setupDataYAxisColumns()


if __name__ == '__main__':
    pass
